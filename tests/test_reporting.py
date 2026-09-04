from io import BytesIO

from crew_compliance.reporting.export import DISCLAIMER, export_csv, export_xlsx
from crew_compliance.reporting.pdf import export_pdf
from crew_compliance.engine.runner import run_analysis
from tests.helpers import make_duty, make_roster


def test_export_contains_disclaimer_and_version():
    result = run_analysis(make_roster([make_duty()]), "easa")
    csv_text = export_csv(result).decode("utf-8")
    assert DISCLAIMER[:40] in csv_text
    assert "easa-ftl-v1" in csv_text
    xlsx = export_xlsx(result)
    assert xlsx[:2] == b"PK"


def test_pdf_export_has_summary_and_disclaimer():
    result = run_analysis(make_roster([make_duty()]), "easa")
    pdf = export_pdf(result, company_name="Northwind Air")
    assert pdf.startswith(b"%PDF")
    assert b"Northwind Air" in pdf
    assert b"Executive summary" in pdf
    assert b"Insufficient-data" in pdf or b"Insufficient" in pdf


def test_upload_templates_have_suggested_headers():
    from openpyxl import load_workbook

    from crew_compliance.reporting.templates import (
        CREDENTIAL_TEMPLATE_HEADERS,
        OPENING_TEMPLATE_HEADERS,
        credential_template_xlsx,
        opening_balance_template_xlsx,
    )

    opening = load_workbook(BytesIO(opening_balance_template_xlsx()))
    creds = load_workbook(BytesIO(credential_template_xlsx()))
    opening_headers = [cell.value for cell in next(opening.active.iter_rows(min_row=1, max_row=1))]
    cred_headers = [cell.value for cell in next(creds.active.iter_rows(min_row=1, max_row=1))]
    assert tuple(opening_headers) == OPENING_TEMPLATE_HEADERS
    assert tuple(cred_headers) == CREDENTIAL_TEMPLATE_HEADERS
    assert opening.active.max_row == 1
    assert creds.active.max_row == 1

from __future__ import annotations

import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from crew_compliance.domain.models import AnalysisResult, Finding

DISCLAIMER = (
    "This report is a compliance screening and review aid. It is not an authoritative "
    "replacement for an operator's approved compliance monitoring system, not legal advice, "
    "and not a determination by a regulator. Findings are potential issues or data limitations "
    "that require review by a qualified aviation professional."
)


def findings_frame(result: AnalysisResult) -> pd.DataFrame:
    rows = [_finding_row(f) for f in result.findings]
    return pd.DataFrame(rows)


def _finding_row(finding: Finding) -> dict:
    return {
        "finding_id": finding.finding_id,
        "kind": finding.kind.value,
        "severity": finding.severity.value,
        "crew_id": finding.crew_id,
        "crew_name": finding.crew_name,
        "date": finding.event_time.date().isoformat() if finding.event_time else "",
        "rule_id": finding.rule_id,
        "rule_name": finding.rule_name,
        "citation": finding.citation,
        "actual": finding.actual,
        "required": finding.required,
        "difference": finding.difference,
        "units": finding.units,
        "framework": finding.framework_id,
        "ruleset_version": finding.ruleset_version,
        "duty_id": finding.duty_id or "",
        "flight_id": finding.flight_id or "",
        "explanation": finding.explanation,
        "assumptions": " | ".join(finding.assumptions),
        "limitations": " | ".join(finding.limitations),
        "opening_balance_status": finding.evidence.get("opening_balance_status", ""),
        "opening_balance_hours": finding.evidence.get("opening_balance_hours", ""),
        "opening_balance_as_of": finding.evidence.get("opening_balance_as_of", ""),
    }


def export_csv(result: AnalysisResult) -> bytes:
    frame = findings_frame(result)
    header = _summary_lines(result)
    buf = io.StringIO()
    for line in header:
        buf.write(f"# {line}\n")
    frame.to_csv(buf, index=False)
    return buf.getvalue().encode("utf-8")


def export_xlsx(result: AnalysisResult) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    fill = PatternFill("solid", fgColor="1F3A5F")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    summary["A1"] = "Crew Compliance Checker — Screening Report"
    summary["A1"].font = title_font
    summary["A1"].fill = fill
    summary.merge_cells("A1:B1")
    for idx, line in enumerate(_summary_lines(result), start=3):
        summary[f"A{idx}"] = line
        summary[f"A{idx}"].alignment = Alignment(wrap_text=True)
    summary.column_dimensions["A"].width = 100

    findings_sheet = workbook.create_sheet("Findings")
    frame = findings_frame(result)
    if frame.empty:
        findings_sheet["A1"] = "No findings."
    else:
        for r_idx, row in enumerate(dataframe_to_rows(frame, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                findings_sheet.cell(r_idx, c_idx, value)

    assumptions = workbook.create_sheet("Assumptions")
    assumptions["A1"] = "Assumptions"
    for i, item in enumerate(result.assumptions, start=2):
        assumptions[f"A{i}"] = item
    limitations = workbook.create_sheet("Limitations")
    limitations["A1"] = "Limitations"
    for i, item in enumerate(result.limitations, start=2):
        limitations[f"A{i}"] = item

    buf = io.BytesIO()
    workbook.save(buf)
    return buf.getvalue()


def _summary_lines(result: AnalysisResult) -> list[str]:
    counts = result.counts_by_severity()
    return [
        DISCLAIMER,
        f"Analyzed at (UTC naive local stamp): {result.analyzed_at.isoformat(timespec='seconds')}",
        f"Roster: {result.source_name}",
        f"Framework: {result.framework_name} ({result.framework_id})",
        f"Ruleset: {result.ruleset_id} version {result.ruleset_version}",
        f"Crew reviewed: {result.crew_reviewed}",
        f"Duties analyzed: {result.duties_analyzed}",
        f"Flights analyzed: {result.flights_analyzed}",
        f"Potential findings: {result.potential_issue_count()}",
        f"Insufficient-data notices: {result.insufficient_data_count()}",
        (
            f"Critical: {counts['critical']}  High: {counts['high']}  "
            f"Medium: {counts['medium']}  Low: {counts['low']}"
        ),
    ]
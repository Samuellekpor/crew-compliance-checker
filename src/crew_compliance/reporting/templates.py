from __future__ import annotations

import io

from openpyxl import Workbook

OPENING_TEMPLATE_HEADERS = (
    "crew_id",
    "crew_name",
    "role",
    "window_type",
    "hours_already_accrued",
    "as_of_date",
)

CREDENTIAL_TEMPLATE_HEADERS = (
    "crew_id",
    "crew_name",
    "role",
    "credential_type",
    "credential_detail",
    "expiry_date",
)


def empty_header_xlsx(headers: tuple[str, ...], sheet_name: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for index, header in enumerate(headers, start=1):
        sheet.cell(1, index, header)
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = max(18, len(header) + 4)
    buf = io.BytesIO()
    workbook.save(buf)
    return buf.getvalue()


def opening_balance_template_xlsx() -> bytes:
    return empty_header_xlsx(OPENING_TEMPLATE_HEADERS, "Opening balances")


def credential_template_xlsx() -> bytes:
    return empty_header_xlsx(CREDENTIAL_TEMPLATE_HEADERS, "Credentials")

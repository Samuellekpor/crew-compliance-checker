from __future__ import annotations

import io
from pathlib import Path
from typing import BinaryIO

import pandas as pd
from openpyxl import load_workbook

from crew_compliance.ingestion.common import MAX_ROWS, MAX_UPLOAD_BYTES, IngestError


def _ensure_size(data: bytes) -> None:
    if len(data) > MAX_UPLOAD_BYTES:
        raise IngestError(
            f"The uploaded file is larger than {MAX_UPLOAD_BYTES // (1024 * 1024)} MB. "
            "Reduce the file size and try again."
        )


def load_table(source: str | Path | BinaryIO | bytes, filename: str | None = None) -> pd.DataFrame:
    if hasattr(source, "read"):
        data = source.read()
        name = filename or getattr(source, "name", "upload")
    elif isinstance(source, (bytes, bytearray)):
        data = bytes(source)
        name = filename or "upload"
    else:
        path = Path(source)
        data = path.read_bytes()
        name = filename or path.name
    _ensure_size(data)
    lower = name.lower()
    if lower.endswith(".csv"):
        return _read_csv(data)
    if lower.endswith(".xlsx"):
        return _read_xlsx(data)
    raise IngestError("Upload a CSV or XLSX file. Other formats are not supported in this version.")


def _read_csv(data: bytes) -> pd.DataFrame:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            text = None
    else:
        raise IngestError("The CSV file could not be read. Save it as UTF-8 and try again.")
    df = pd.read_csv(io.StringIO(text))
    if len(df) > MAX_ROWS:
        raise IngestError(f"The file has more than {MAX_ROWS:,} rows. Split the roster and try again.")
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _read_xlsx(data: bytes) -> pd.DataFrame:
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True, keep_vba=False)
    except Exception as exc:
        raise IngestError("The spreadsheet could not be opened. Export a simple .xlsx without macros.") from exc
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        raise IngestError("The spreadsheet is empty.")
    headers = [str(h).strip() if h is not None else f"column_{i}" for i, h in enumerate(rows[0])]
    body = rows[1:]
    if len(body) > MAX_ROWS:
        raise IngestError(f"The file has more than {MAX_ROWS:,} rows. Split the roster and try again.")
    return pd.DataFrame(body, columns=headers)